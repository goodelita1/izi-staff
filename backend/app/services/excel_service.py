from datetime import date, datetime
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from app.models.inventory import Inventory

COLUMNS = [
    ("inventory_number", "Інвентарний №"),
    ("invoice_number", "Номер накладної"),
    ("invoice_date", "Дата накладної"),
    ("item_name", "Назва майна"),
    ("nomenclature_code", "Код номенклатури"),
    ("serial_number", "Серійний номер"),
    ("unit", "Одиниця"),
    ("category", "Категорія"),
    ("quantity", "Кількість"),
    ("price", "Вартість"),
    ("total_price", "Сума"),
    ("issued_to", "Кому видано"),
    ("location", "Місце"),
    ("issued_date", "Дата видачі"),
    ("ownership", "Тип власності"),
    ("department", "Служба"),
    ("invoice_receiver", "Отримувач"),
    ("status", "Статус"),
    ("note", "Примітка"),
]

VALID_STATUSES = {"Warehouse", "Issued", "Returned", "Written Off"}
VALID_OWNERSHIPS = {"Own", "State"}

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="F9FAFB")
COL_WIDTHS = {
    "inventory_number": 20,
    "invoice_number": 20,
    "invoice_date": 14,
    "item_name": 30,
    "nomenclature_code": 18,
    "serial_number": 20,
    "unit": 10,
    "category": 16,
    "quantity": 10,
    "price": 12,
    "total_price": 12,
    "issued_to": 20,
    "location": 18,
    "issued_date": 14,
    "ownership": 14,
    "department": 16,
    "invoice_receiver": 20,
    "status": 14,
    "note": 24,
}


def _parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


class ImportError(Exception):
    def __init__(self, row: int, field: str, message: str) -> None:
        self.row = row
        self.field = field
        self.message = message
        super().__init__(f"Row {row}, {field}: {message}")


class ExcelService:
    def export(self, items: list[Inventory]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Інвентар"

        for col_idx, (field, header) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[cell.column_letter].width = COL_WIDTHS.get(field, 18)

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        for row_idx, item in enumerate(items, start=2):
            for col_idx, (field, _) in enumerate(COLUMNS, start=1):
                value = getattr(item, field, None)
                if isinstance(value, (date, datetime)):
                    value = value.isoformat()
                ws.cell(row=row_idx, column=col_idx, value=value)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def import_file(self, file_bytes: bytes) -> tuple[list[dict], list[dict]]:
        """Returns (valid_records, errors). Errors: [{row, field, message}]."""
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active

        raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        field_map = {header: field for field, header in COLUMNS}

        records: list[dict] = []
        errors: list[dict] = []

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(row):
                continue

            record: dict = {}
            for col_idx, value in enumerate(row):
                header = raw_headers[col_idx] if col_idx < len(raw_headers) else None
                if header and header in field_map:
                    record[field_map[header]] = value

            row_errors = self._validate_row(row_idx, record)
            if row_errors:
                errors.extend(row_errors)
            else:
                record = self._coerce_row(record)
                records.append(record)

        return records, errors

    def _validate_row(self, row: int, record: dict) -> list[dict]:
        errs = []

        if not record.get("item_name"):
            errs.append(
                {"row": row, "field": "item_name", "message": "Назва майна обов'язкова"}
            )
        if not record.get("invoice_number"):
            errs.append(
                {
                    "row": row,
                    "field": "invoice_number",
                    "message": "Номер накладної обов'язковий",
                }
            )
        if _parse_date(record.get("invoice_date")) is None:
            errs.append(
                {
                    "row": row,
                    "field": "invoice_date",
                    "message": "Дата накладної обов'язкова або невірний формат",
                }
            )

        qty = record.get("quantity")
        if qty is not None:
            try:
                if int(qty) < 1:
                    errs.append(
                        {
                            "row": row,
                            "field": "quantity",
                            "message": "Кількість мінімум 1",
                        }
                    )
            except (TypeError, ValueError):
                errs.append(
                    {
                        "row": row,
                        "field": "quantity",
                        "message": "Кількість має бути числом",
                    }
                )

        for field in ("price", "total_price"):
            val = record.get(field)
            if val is not None:
                try:
                    if float(val) < 0:
                        errs.append(
                            {
                                "row": row,
                                "field": field,
                                "message": "Значення не може бути від'ємним",
                            }
                        )
                except (TypeError, ValueError):
                    errs.append(
                        {"row": row, "field": field, "message": "Має бути числом"}
                    )

        status = record.get("status")
        if status and status not in VALID_STATUSES:
            errs.append(
                {
                    "row": row,
                    "field": "status",
                    "message": f"Допустимі значення: {', '.join(VALID_STATUSES)}",
                }
            )

        ownership = record.get("ownership")
        if ownership and ownership not in VALID_OWNERSHIPS:
            errs.append(
                {
                    "row": row,
                    "field": "ownership",
                    "message": f"Допустимі значення: {', '.join(VALID_OWNERSHIPS)}",
                }
            )

        return errs

    def _coerce_row(self, record: dict) -> dict:
        record.pop("inventory_number", None)
        record["invoice_date"] = _parse_date(record.get("invoice_date"))
        record["issued_date"] = _parse_date(record.get("issued_date"))
        record["quantity"] = int(record.get("quantity") or 1)
        record["price"] = float(record.get("price") or 0.0)
        record["total_price"] = float(record.get("total_price") or 0.0)
        record.setdefault("unit", "шт")
        record.setdefault("category", "Інше")
        record.setdefault("status", "Warehouse")
        return record
